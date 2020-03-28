# Import Package
import openpyxl

# Load Workbook
# wk = openpyxl.load_workbook("D:\\TestData.xlsx")
wk = openpyxl.load_workbook("C:\\Users\\mayank.saxena2\\PycharmProjects\\Demo1\\TestData\\TestData.xlsx")

def fetch_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = wk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value

# print(fetch_number_of_rows('Login'))
# print(fetch_cell_data('Login', 2, 2))

# print(wk.sheetnames)
# print("Active sheet is: "+wk.active.title)

# Create object of any sheet on which you want to work
# sh = wk['Login']

# print(sh.title)
# print(sh['A1'].value)
# print(sh['B2'].value)
# c1 = sh.cell(1,2)
# print(c1.value)
# print(c1.row)
# print(c1.column)

# Find Rows and Columns having data
# rows = sh.max_row
# columns = sh.max_column

# print("Total rows are: "+str(rows))
# print("Total columns are: "+str(columns))

# for i in range(1,rows+1):
#    for j in range(1,columns+1):
#        c=sh.cell(i,j)
#        print(c.value)